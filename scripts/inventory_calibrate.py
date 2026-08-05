#!/usr/bin/env python3
"""用 I-485 库存的「按 PD 年份分桶」实测，校准模型近端(当前 cutoff 正前方)的需求密度。

为什么只校准近端、不碰 f：
  库存只统计「已递交 I-485」的人。EB-1 中国只有 PD 已 current(或进入表B窗口)才能递 I-485，
  所以 2024+ 的 PD 在库存里基本是 0——不是没人，而是还递不了。故库存暂时仍**钉不死** 2024 PD 的 f
  (与 estimate_f 的判断一致)。但它能给出 cutoff **正前方**(≈PD 当前年)的真实"墙"，
  用来校准/验证模型近端的推进速度——这是不引入假精度的、最扎实的一块。

单位对齐：
  模型 density = 主申请人 / PD-日 (simulate: advanceDays = 月签发主申 / density)。
  库存计的是 I-485 申请数(含配偶子女派生)，故 ÷家庭系数 得主申，再 ÷365 得 主申/PD-日。
  注意：库存是「剩余仍在排队」的快照——cutoff 已越过较久的 PD 年大多已清空(非原始墙)，
  只有 ≈cutoff 当年的桶最具代表性。库存只含已递件者，是真实墙的**下界**。

用法：python scripts/inventory_calibrate.py
输出写 stdout + (若设)GITHUB_STEP_SUMMARY。
"""
import glob
import os
import re
from datetime import date

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
USCIS = os.path.join(ROOT, "data/raw/uscis")
INDEX = os.path.join(ROOT, "index.html")
MONTHS = ['january', 'february', 'march', 'april', 'may', 'june',
          'july', 'august', 'september', 'october', 'november', 'december']


def newest_inventory_file():
    """按 (年,月) 取最新的 I-485 库存 xlsx 路径。"""
    best = None
    for f in glob.glob(os.path.join(USCIS, "I485_Pending_Inventory_*.xlsx")):
        m = re.search(r"I485_Pending_Inventory_([a-z]+)_(\d{4})", os.path.basename(f))
        if m and m.group(1) in MONTHS:
            ym = (int(m.group(2)), MONTHS.index(m.group(1)) + 1)
            if best is None or ym > best[0]:
                best = (ym, f)
    return best  # ((year,month), path) or None


def eb1_china_by_pd_year(path):
    """读 China sheet，返回 {pd_year_label: (确切计数, 被抑制D的格数)}（EB1，Available+Awaiting 合计）。
    表头形如 'Priority Date Year - 2023' / '... - Prior Years'。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "China" not in wb.sheetnames:
        return {}, None
    rows = list(wb["China"].iter_rows(values_only=True))
    # 真表头行有"多"个 'Priority Date Year - XXXX' 列(标题行只在一句话里出现一次)；取出现最多的行
    def n_year_cells(r):
        return sum(1 for c in (r or []) if c and 'Priority Date Year' in str(c))
    hdr_idx = max(range(min(8, len(rows))), key=lambda i: n_year_cells(rows[i]), default=None)
    if hdr_idx is None or n_year_cells(rows[hdr_idx]) < 3:
        return {}, None
    hdr = rows[hdr_idx]
    # 列 -> 年标签('Priority Date Year - 2023' → '2023'；'... - Prior Years' → 'Prior Years')
    col_year = {}
    for ci, c in enumerate(hdr):
        if c and 'Priority Date Year' in str(c):
            col_year[ci] = str(c).split(' - ', 1)[-1].strip()
    asof = None
    for r in rows[:hdr_idx]:
        for c in (r or []):
            m = re.search(r'As of (\w+)\s+\d+,?\s*(\d{4})', str(c or ''))
            if m:
                asof = f"{m.group(2)}-{m.group(1)[:3]}"
    out = {}
    for r in rows:
        if not r or not str(r[1] or '').startswith('Employment-Based 1'):
            continue
        for ci, lab in col_year.items():
            v = r[ci] if ci < len(r) else None
            cnt, dn = out.get(lab, (0, 0))
            if isinstance(v, (int, float)):
                cnt += int(v)
            elif str(v).strip() == 'D':
                dn += 1
            out[lab] = (cnt, dn)
    return out, asof


def model_near_density():
    """从 index.html PRESETS.realistic 取近端密度 densityHigh(=PD≈2023 的 主申/PD-日)、family，
    及当前 表A cutoff / 表B DFF。正则失败时打印警告再用默认值——静默回退曾导致与线上真值脱节。"""
    s = open(INDEX, encoding="utf-8").read()
    m = re.search(r"realistic:\s*\{[^}]*?familyMultiplier:([\d.]+)[^}]*?densityHigh:([\d.]+)", s)
    if not m:
        print("[warn] 未能从 index.html 读到 familyMultiplier/densityHigh，使用默认 1.9/8.0")
    fam = float(m.group(1)) if m else 1.9
    dh = float(m.group(2)) if m else 8.0
    cm = re.search(r"'EB-1A':\s*\{\s*'CN':\s*\{\s*A:\s*'([0-9-]+)',\s*B:\s*'([0-9-]+)'", s)
    if not cm:
        print("[warn] 未能从 index.html 读到 表A/表B cutoff，使用默认 2023-06-01/2023-12-01")
    cutoff = date.fromisoformat(cm.group(1)) if cm else date(2023, 6, 1)
    dff = date.fromisoformat(cm.group(2)) if cm else date(2023, 12, 1)
    return fam, dh, cutoff, dff


def main():
    L = []
    def p(s=""):
        print(s); L.append(str(s))

    inv = newest_inventory_file()
    if not inv:
        p("## 库存近端校准：未找到 I-485 库存文件")
        return
    (iy, im), path = inv
    dist, asof = eb1_china_by_pd_year(path)
    fam, dens_model, cutoff, dff = model_near_density()

    p(f"## I-485 库存近端校准 (EB-1 中国, 库存 {iy}-{im:02d}, as of {asof or '?'})")
    p("PD年 | I-485在案(含派生) | 被抑制格 | ≈主申(÷family) ")
    p("|---|---|---|---|")
    for lab in ['Prior Years', '2020', '2021', '2022', '2023', '2024', '2025', '2026']:
        if lab not in dist:
            continue
        cnt, dn = dist[lab]
        prin = cnt / fam
        p(f"| {lab} | {cnt:,} | D×{dn} | {prin:,.0f} |")

    near = str(cutoff.year)
    if near in dist and dist[near][0] > 0:
        cnt = dist[near][0]
        prin = cnt / fam
        # 分母 = 该桶实际可递件的 PD 跨度，而非全年 365 天：
        # 只有 PD ≤ 表A cutoff(或表B窗口开放期 ≤ DFF)的人才能递 I-485，
        # 故 PD-{near} 桶里的库存集中在 [年初, max(表A,表B)] 区间。
        # 误用 365 会按 ~1.1–2.4× 稀释实测密度，把下界报松、掩盖模型偏乐观。
        year_start = date(cutoff.year, 1, 1)
        span_end = min(max(cutoff, dff), date(cutoff.year, 12, 31))
        span_days = max(1, (span_end - year_start).days)
        dens_inv = prin / span_days     # 主申/PD-日(下界:仅已递件、且为剩余快照)
        p("")
        p(f"### 近端密度校准（cutoff={cutoff}，对照 PD-{near} 桶）")
        p(f"- 库存实测(下界)：PD-{near} 在案 {cnt:,} 件 → ≈{prin:,.0f} 主申 ÷ 可递件跨度 "
          f"{span_days} 天（年初→max(表A {cutoff}, 表B {dff})）→ **{dens_inv:.1f} 主申/PD-日**")
        p(f"- 模型现行(densityHigh)：**{dens_model:.1f} 主申/PD-日**")
        # 2026-08 重标定后，densityHigh 是「真实排队密度」(含 PD 未 current、尚不能递 I-485 的人)，
        # 而库存实测只含「已递件」者。二者的预期关系不是"越高越保守"，而是结构性倍数：
        # 待签池/I-485库存 原始比 ≈ 2.45，故合理区间取 ×1.8–3.0。
        ratio = dens_model / dens_inv if dens_inv else 0
        LO, HI = 1.8, 3.0
        if LO <= ratio <= HI:
            p(f"- 判定：✅ 模型({dens_model:.1f}) / 实测下界({dens_inv:.1f}) = **×{ratio:.2f}**，"
              f"落在预期区间 ×{LO}–{HI}（待签池/库存原始比 ≈2.45）。真实密度含未递件者，故应高于下界，量级吻合。")
        elif ratio < LO:
            p(f"- 判定：⚠️ 倍数偏低（×{ratio:.2f} < {LO}）：模型密度可能低估了「尚未递 I-485」的隐藏排队，"
              f"近端推进偏快(乐观)。参考上调区间 {dens_inv*LO:.0f}–{dens_inv*HI:.0f}。")
        else:
            p(f"- 判定：⚠️ 倍数偏高（×{ratio:.2f} > {HI}）：模型密度可能高估隐藏排队，近端推进偏慢(悲观)。"
              f"参考下调区间 {dens_inv*LO:.0f}–{dens_inv*HI:.0f}。")
        p(f"- 注：供给与密度须同口径同步——现行总供给 ≈6,983 张/年（对齐实际发卡 6,980），"
          f"密度 {dens_model:.1f} 人/天；二者比值决定推进速度，单独改一个会破坏标定。")
        p("- 注：库存为「剩余排队」快照，仅 ≈cutoff 当年桶具代表性；2024+ 桶=0 系"
          "「PD 未 current、尚不能递 I-485」，故 f 仍不可由库存钉死(见 estimate_f)。")
    else:
        p(f"\nPD-{near} 桶为空或缺失：cutoff 当年尚无已递件可测，跳过近端校准。")

    sp = os.environ.get("GITHUB_STEP_SUMMARY")
    if sp:
        try:
            with open(sp, "a", encoding="utf-8") as fh:
                fh.write("\n".join(L) + "\n")
        except OSError:
            pass


if __name__ == "__main__":
    main()
