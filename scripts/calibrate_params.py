#!/usr/bin/env python3
"""
Calibrate V21 predictor parameters from actual USCIS/DOS data.

Analyzes:
  1. DOS Monthly IV Issuance → total China EB-1 supply, spillover decomposition
  2. I-485 Pending Inventory → PD density (applicants per day)
  3. Cross-reference → family multiplier
"""
import os
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
DOS_DIR = REPO / "data" / "raw" / "dos"
USCIS_DIR = REPO / "data" / "raw" / "uscis"

MONTHS_ORDER = [
    "october", "november", "december",
    "january", "february", "march", "april", "may",
    "june", "july", "august", "september",
]
CAL_MONTHS = [
    "january", "february", "march", "april", "may", "june",
    "july", "august", "september", "october", "november", "december",
]

INA_GLOBAL_QUOTA = 140_000
EB1_SHARE = 0.286
CHINA_CAP = 0.07
BASE_QUOTA = round(INA_GLOBAL_QUOTA * EB1_SHARE * CHINA_CAP)  # 2803


# ─── DOS Monthly Issuance ───────────────────────────────────────────

def parse_dos_file(path):
    """Extract visa issuances by (country_group, visa_class) from a DOS XLSX."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    data = {}
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        country = str(row[0]).strip()
        vclass = str(row[1]).strip()
        count = row[2] if row[2] else 0
        if not isinstance(count, (int, float)):
            continue

        if "China - mainland" in country:
            grp = "China"
        elif "India" in country:
            grp = "India"
        else:
            grp = "ROW"

        key = (grp, vclass)
        data[key] = data.get(key, 0) + int(count)
    wb.close()
    return data


def fiscal_year_from_filename(fname):
    """iv_issuance_october_2024.xlsx → (FY2025, 'october', 2024)"""
    m = re.match(r"iv_issuance_(\w+)_(\d{4})\.xlsx", fname)
    if not m:
        return None
    month, year = m.group(1), int(m.group(2))
    if month in ("october", "november", "december"):
        fy = year + 1
    else:
        fy = year
    return fy, month, year


def analyze_dos():
    """Aggregate DOS data by fiscal year."""
    fy_data = defaultdict(lambda: defaultdict(int))
    monthly_china_eb1 = {}

    for f in sorted(DOS_DIR.glob("iv_issuance_*.xlsx")):
        info = fiscal_year_from_filename(f.name)
        if not info:
            continue
        fy, month, cal_year = info
        data = parse_dos_file(f)

        label = f"{month.title()} {cal_year}"
        china_e1 = data.get(("China", "E1"), 0)
        india_e1 = data.get(("India", "E1"), 0)
        monthly_china_eb1[label] = china_e1

        for (grp, vclass), cnt in data.items():
            fy_data[fy][(grp, vclass)] += cnt

    return fy_data, monthly_china_eb1


def decompose_supply(fy_data):
    """For each FY, calculate supply components."""
    results = {}
    for fy in sorted(fy_data.keys()):
        d = fy_data[fy]

        china_eb1 = d.get(("China", "E1"), 0)
        india_eb1 = d.get(("India", "E1"), 0)
        row_eb1 = d.get(("ROW", "E1"), 0)

        total_eb1 = china_eb1 + india_eb1 + row_eb1

        india_quota = BASE_QUOTA  # India also gets 7% = 2803
        row_quota = INA_GLOBAL_QUOTA * EB1_SHARE - BASE_QUOTA - india_quota
        # ROW quota = 40040 - 2803 - 2803 = 34434 (shared among all other countries)

        # EB-4/5 usage
        eb4_classes = ["SD", "SE", "SF", "SG", "SH", "SI", "SJ", "SK", "SR"]
        eb5_classes = ["C5", "I5", "R5", "T5"]
        eb4_total = sum(d.get((g, c), 0) for g in ["China", "India", "ROW"] for c in eb4_classes)
        eb5_total = sum(d.get((g, c), 0) for g in ["China", "India", "ROW"] for c in eb5_classes)

        eb4_quota = round(INA_GLOBAL_QUOTA * 0.071)  # ~9940
        eb5_quota = round(INA_GLOBAL_QUOTA * 0.071)

        eb4_unused = max(0, eb4_quota - eb4_total)
        eb5_unused = max(0, eb5_quota - eb5_total)
        eb45_spillover_pool = eb4_unused + eb5_unused

        # ROW EB-1 unused (spillover to China/India)
        row_unused = max(0, row_quota - row_eb1)

        # What China actually got beyond base quota
        china_excess = china_eb1 - BASE_QUOTA

        results[fy] = {
            "china_eb1": china_eb1,
            "india_eb1": india_eb1,
            "row_eb1": row_eb1,
            "total_eb1": total_eb1,
            "eb4_total": eb4_total,
            "eb5_total": eb5_total,
            "eb4_unused": eb4_unused,
            "eb5_unused": eb5_unused,
            "eb45_spillover_pool": eb45_spillover_pool,
            "row_unused": row_unused,
            "china_excess": china_excess,
        }
    return results


# ─── I-485 Inventory ────────────────────────────────────────────────

def parse_i485_inventory(path):
    """Parse I-485 inventory XLSX to get China EB-1 counts by PD month."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    china_sheet = None
    for name in wb.sheetnames:
        if "china" in name.lower() or "mainland" in name.lower():
            china_sheet = name
            break

    if not china_sheet:
        for name in wb.sheetnames:
            if name != "Grand Totals" and name != "Sheet1":
                china_sheet = name
                break

    if not china_sheet:
        wb.close()
        return {}

    ws = wb[china_sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return {}

    # Find header row and EB-1 column
    header = None
    header_idx = 0
    for i, row in enumerate(rows):
        if row and any(str(c).strip().upper().startswith("EB") or "1ST" in str(c).upper() for c in row if c):
            header = row
            header_idx = i
            break

    if not header:
        return {}

    # Find EB-1 column index
    eb1_col = None
    for j, cell in enumerate(header):
        s = str(cell).strip().upper() if cell else ""
        if "1ST" in s or s == "EB-1" or s == "EB1" or "FIRST" in s:
            eb1_col = j
            break

    if eb1_col is None:
        return {}

    # Parse PD rows: column 0 is usually the PD date/period
    density = {}
    for row in rows[header_idx + 1:]:
        if not row or not row[0]:
            continue
        pd_label = str(row[0]).strip()
        count = row[eb1_col] if eb1_col < len(row) else None

        if count is None or count == "" or count == "D" or count == "N/A":
            continue
        try:
            count = int(float(str(count).replace(",", "")))
        except (ValueError, TypeError):
            continue

        density[pd_label] = count

    return density


def analyze_i485():
    """Analyze all I-485 inventory files for PD density."""
    all_snapshots = {}

    for f in sorted(USCIS_DIR.glob("I485_Pending_Inventory_*.xlsx")):
        density = parse_i485_inventory(f)
        if density:
            all_snapshots[f.stem] = density

    return all_snapshots


# ─── Main Analysis ──────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("EB1A V21 参数校准 — 基于实际 USCIS/DOS 数据")
    print("=" * 70)

    # ── 1. DOS 签发数据分析 ──
    print("\n" + "─" * 70)
    print("1. DOS 月度签发数据 (实际 supply)")
    print("─" * 70)

    fy_data, monthly = analyze_dos()
    supply = decompose_supply(fy_data)

    for fy in sorted(supply.keys()):
        s = supply[fy]
        months_count = sum(1 for f in DOS_DIR.glob("iv_issuance_*.xlsx")
                          if fiscal_year_from_filename(f.name) and fiscal_year_from_filename(f.name)[0] == fy)
        print(f"\n  FY{fy} ({months_count} months of data):")
        print(f"    China EB-1 签发:    {s['china_eb1']:>6}")
        print(f"    India EB-1 签发:    {s['india_eb1']:>6}")
        print(f"    ROW EB-1 签发:      {s['row_eb1']:>6}")
        print(f"    EB-1 全球总计:      {s['total_eb1']:>6}")
        print(f"    EB-4 全球使用:      {s['eb4_total']:>6} (配额 ~9,940)")
        print(f"    EB-5 全球使用:      {s['eb5_total']:>6} (配额 ~9,940)")
        print(f"    EB-4 未用 (→溢出):  {s['eb4_unused']:>6}")
        print(f"    EB-5 未用 (→溢出):  {s['eb5_unused']:>6}")
        print(f"    ROW EB-1 余量:      {s['row_unused']:>6}")
        print(f"    China 超出基础配额: {s['china_excess']:>+6} (= {s['china_eb1']} - {BASE_QUOTA})")

    # Monthly breakdown
    print(f"\n  月度 China EB-1 签发:")
    for label, count in sorted(monthly.items(), key=lambda x: x[0]):
        bar = "█" * (count // 50)
        print(f"    {label:<20} {count:>5}  {bar}")

    # ── 2. Supply 参数推导 ──
    print("\n" + "─" * 70)
    print("2. Supply 参数推导")
    print("─" * 70)

    for fy in sorted(supply.keys()):
        s = supply[fy]
        total_china = s["china_eb1"]
        excess = s["china_excess"]

        # Decomposition logic:
        # China gets: base(2803) + share_of_ROW_unused + share_of_eb45_unused + india_effect
        # india_effect: if India uses less than quota, excess goes to China (positive)
        #               if India uses more than quota, it competes (but EB-1 India quota is same)

        india_excess = s["india_eb1"] - BASE_QUOTA  # India's usage above its 7% base

        # EB-4/5 spillover to EB-1 (INA: unused EB-4/5 → EB-1 first)
        # Only a fraction goes to China (China's share of oversubscribed demand)
        eb45_pool = s["eb45_spillover_pool"]

        # ROW unused EB-1 is split between China and India (by demand)
        # Approximate: China gets proportional share
        china_india_demand = s["china_eb1"] + s["india_eb1"]
        if china_india_demand > 0:
            china_share_of_row = round(s["row_unused"] * s["china_eb1"] / china_india_demand)
        else:
            china_share_of_row = 0

        # What's left = eb45 spillover to China
        implied_eb45_to_china = excess - china_share_of_row
        if implied_eb45_to_china < 0:
            china_share_of_row = excess
            implied_eb45_to_china = 0

        print(f"\n  FY{fy} 分解:")
        print(f"    China EB-1 总签发:         {total_china}")
        print(f"    - 基础配额:                {BASE_QUOTA}")
        print(f"    = 超额部分:                {excess}")
        print(f"    其中:")
        print(f"      ROW 溢出 (China 份额):   ~{china_share_of_row}")
        print(f"      EB-4/5 溢出 (China):     ~{implied_eb45_to_china}")
        print(f"      (EB-4/5 全球未用池:      {eb45_pool})")
        print(f"    India EB-1 超额:           {india_excess:+d} (>0 = 印度抢占)")

    # ── 3. I-485 Inventory 密度分析 ──
    print("\n" + "─" * 70)
    print("3. I-485 Pending Inventory — PD 密度分析")
    print("─" * 70)

    snapshots = analyze_i485()

    if not snapshots:
        print("  (无法解析 I-485 inventory 文件)")
    else:
        # Use the latest snapshot
        latest_key = sorted(snapshots.keys())[-1]
        latest = snapshots[latest_key]
        print(f"\n  最新快照: {latest_key}")
        print(f"  PD 区间数: {len(latest)}")

        # Sort by PD and show
        print(f"\n  PD 月份 → 待审人数 (China EB-1):")
        total_pending = 0
        density_2023_2024 = []
        density_2024_plus = []

        for pd, count in sorted(latest.items()):
            total_pending += count
            bar = "█" * (count // 20)
            print(f"    {pd:<12} {count:>5}  {bar}")

            # Classify by year for density calculation
            try:
                if "2023" in pd or "2024" in pd:
                    density_2023_2024.append(count)
                if "2024" in pd or "2025" in pd or "2026" in pd:
                    density_2024_plus.append(count)
            except:
                pass

        print(f"\n  总待审 (China EB-1): {total_pending} 主申请人")

        if density_2023_2024:
            avg_per_month_2324 = sum(density_2023_2024) / len(density_2023_2024)
            per_day_2324 = avg_per_month_2324 / 30.44
            print(f"\n  2023-2024 PD 密度:")
            print(f"    月均: {avg_per_month_2324:.0f} 主申/月")
            print(f"    日均: {per_day_2324:.1f} 主申/天")

        if density_2024_plus:
            avg_per_month_24p = sum(density_2024_plus) / len(density_2024_plus)
            per_day_24p = avg_per_month_24p / 30.44
            print(f"\n  2024+ PD 密度:")
            print(f"    月均: {avg_per_month_24p:.0f} 主申/月")
            print(f"    日均: {per_day_24p:.1f} 主申/天")

    # ── 4. Family Multiplier ──
    print("\n" + "─" * 70)
    print("4. 家庭系数 (family multiplier) 推导")
    print("─" * 70)

    if snapshots and supply:
        latest_data = snapshots[sorted(snapshots.keys())[-1]]
        total_principals = sum(latest_data.values())

        # Use FY2025 total China EB-1 visas
        fy25 = supply.get(2025, {})
        fy24 = supply.get(2024, {})

        if fy25:
            print(f"\n  FY2025 China EB-1 签证总数: {fy25['china_eb1']}")
            print(f"  I-485 Inventory 总主申请人:  {total_principals}")
            print(f"  (注: 签证数包含家属, inventory 只计主申)")

            # Historical approach: visas_issued / principals_served
            # In one year, roughly china_eb1 visas serve china_eb1/multiplier principals
            # So multiplier = visas / principals_processed
            # But we don't directly know principals_processed per year

            # Better: look at change in inventory between two snapshots
            # Δinventory = new_filings - visas_issued/multiplier
            # visas_issued/multiplier = principals who got green cards

    # ── 5. 推荐参数 ──
    print("\n" + "=" * 70)
    print("5. 推荐参数 (基于数据)")
    print("=" * 70)

    # Calculate recommended values
    fy_list = sorted(supply.keys())
    if len(fy_list) >= 2:
        avg_china = sum(supply[fy]["china_eb1"] for fy in fy_list) / len(fy_list)
        avg_excess = sum(supply[fy]["china_excess"] for fy in fy_list) / len(fy_list)
        avg_india_excess = sum(supply[fy]["india_eb1"] - BASE_QUOTA for fy in fy_list) / len(fy_list)

        # ROW spillover: use the total excess, then decompose
        # For now, we'll use direct measurement
        print(f"\n  两年均值:")
        print(f"    China EB-1 年签发:  {avg_china:.0f}")
        print(f"    超出基础配额:       {avg_excess:+.0f}")

    for fy in fy_list:
        s = supply[fy]
        total_supply = s["china_eb1"]
        print(f"\n  FY{fy} 实际总 supply = {total_supply}")
        print(f"    V21 默认总 supply = {BASE_QUOTA} + 800 + 400 + 200 = {BASE_QUOTA + 800 + 400 + 200}")
        print(f"    差异: {total_supply - (BASE_QUOTA + 800 + 400 + 200):+d}")

    print()
    print("  ┌────────────────────────┬──────────┬──────────┐")
    print("  │ 参数                   │ V21 默认 │ 数据推荐 │")
    print("  ├────────────────────────┼──────────┼──────────┤")

    if len(fy_list) >= 2:
        # spilloverROW: excess - india_component - eb45_component
        rec_total_excess = round(avg_excess)
        # Simple: attribute proportionally
        avg_eb45_pool = sum(supply[fy]["eb45_spillover_pool"] for fy in fy_list) / len(fy_list)
        # China's share of eb45 ~ china_eb1 / (china_eb1 + india_eb1)
        avg_china_share = sum(supply[fy]["china_eb1"] / max(1, supply[fy]["china_eb1"] + supply[fy]["india_eb1"]) for fy in fy_list) / len(fy_list)
        rec_eb45 = round(avg_eb45_pool * avg_china_share / 100) * 100  # round to 100

        rec_india = round(avg_india_excess / 100) * 100
        rec_row = rec_total_excess - rec_eb45 - rec_india
        if rec_row < 0:
            rec_row = round(rec_total_excess * 0.6 / 100) * 100
            rec_eb45 = round(rec_total_excess * 0.2 / 100) * 100
            rec_india = rec_total_excess - rec_row - rec_eb45

        print(f"  │ spilloverROW           │ {800:>8} │ {rec_row:>+8} │")
        print(f"  │ spilloverIndia         │ {400:>8} │ {rec_india:>+8} │")
        print(f"  │ eb4eb5Spillover        │ {200:>8} │ {rec_eb45:>+8} │")
    else:
        print("  │ (DOS 数据不足, 需要 2 个完整 FY)       │")

    if snapshots:
        latest_data = snapshots[sorted(snapshots.keys())[-1]]
        counts_2324 = []
        counts_24p = []
        for pd, count in latest_data.items():
            if "2023" in pd or "2024" in pd:
                counts_2324.append(count)
            if "2024" in pd or "2025" in pd:
                counts_24p.append(count)

        if counts_2324:
            d_high = sum(counts_2324) / len(counts_2324) / 30.44
            print(f"  │ densityHigh (2023-24)  │ {15:>8} │ {d_high:>8.1f} │")
        if counts_24p:
            d_peak = sum(counts_24p) / len(counts_24p) / 30.44
            print(f"  │ densityPeak (2024+)    │ {18:>8} │ {d_peak:>8.1f} │")

    print(f"  │ familyMultiplier       │ {1.9:>8} │   (下文) │")
    print("  └────────────────────────┴──────────┴──────────┘")

    # Family multiplier analysis using inventory delta
    if len(snapshots) >= 2:
        print("\n  家庭系数推导:")
        keys = sorted(snapshots.keys())
        earliest = snapshots[keys[0]]
        latest = snapshots[keys[-1]]
        t_early = sum(earliest.values())
        t_late = sum(latest.values())
        delta = t_late - t_early
        print(f"    最早快照 ({keys[0]}): {t_early} 主申")
        print(f"    最新快照 ({keys[-1]}): {t_late} 主申")
        print(f"    变化: {delta:+d} 主申")

        # In the period between snapshots:
        # new_filings - (visas_issued / multiplier) = delta
        # So: multiplier = visas_issued / (new_filings - delta)
        # We don't know new_filings directly, but we know visas_issued from DOS
        # This is underdetermined with one equation... use reasonable estimate
        print(f"    (需要 new_filings 数据来精确计算 — 可从 I-140 季报推导)")


if __name__ == "__main__":
    main()
