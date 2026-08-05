#!/usr/bin/env python3
"""I-140 趋势报告：China EB-1A/NIW 申请潮 + 待签存量走势。

数据(由 check_data_updates.py 自动抓取，存 data/raw/uscis/)：
  - I140_I360_I526_Approved_FY*_Q*.xlsx : EB1 China awaiting-visa 存量(领先指标，最干净)
  - I140_FY*_Q*.xlsx                    : I-140 收件(Rec-COB) China EB-1A / NIW

用法：python scripts/i140_trend_report.py
输出可读趋势 + (若设 GITHUB_STEP_SUMMARY)写入运行摘要。
"""
import os, re, glob
from datetime import date
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
USCIS = os.path.join(ROOT, "data/raw/uscis")
MO = {m: i + 1 for i, m in enumerate(
    ['January','February','March','April','May','June','July','August','September','October','November','December'])}


def _asof(rows):
    for r in rows[:5]:
        if r and r[0] and 'As of' in str(r[0]):
            m = re.search(r'As of (\w+)\s+(\d+)?,?\s*(\d{4})', str(r[0]))
            if m:
                return date(int(m.group(3)), MO[m.group(1)], int(m.group(2) or 1))
    return None


def awaiting_stock():
    """China EB1 待签存量时间序列。"""
    out = []
    for f in glob.glob(os.path.join(USCIS, "I140_I360_I526_Approved_FY*_Q*.xlsx")):
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]; rows = list(ws.iter_rows(values_only=True))
        asof = _asof(rows); china = None
        for r in rows:
            if r and str(r[0]).strip() == 'China':
                china = r[1]; break
        # openpyxl 可能把整数读成 float——按数值接受，避免静默丢掉某期存量点
        if asof and isinstance(china, (int, float)):
            out.append((asof, int(china)))
    out.sort()
    return out


def _csv_china_cols(path):
    """读 *_rec_cob.csv 的 CHINA 行，返回各列数值列表(索引同 xlsx：[國家,E11,E12,E13,E21,NIW,...])。
    数字带千分逗号+引号，用 csv 模块解析；非数字(国家名)置 None。"""
    import csv
    with open(path, newline='', encoding='utf-8-sig') as fh:
        for row in csv.reader(fh):
            if row and row[0].strip().upper() == 'CHINA':
                def num(x):
                    x = (x or '').replace(',', '').strip()
                    return int(x) if x.lstrip('-').isdigit() else None
                return [num(c) for c in row]
    return None


def receipts():
    """China EB-1A / NIW I-140 收件(Rec-COB)。各文件为单季值(同 FY 内递减即证)。
    数据源优先级：xlsx(Rec-COB) > CSV(*_rec_cob.csv) > supplement.json。E11=EB-1A, 第5列=NIW。"""
    rec = {}  # tag -> (eb1a, niw)
    # 1) xlsx（最权威，多 sheet）
    for f in sorted(glob.glob(os.path.join(USCIS, "I140_FY*_Q*.xlsx"))):
        m = re.search(r'I140_(FY\d+_Q\d+)', os.path.basename(f))
        if not m:
            continue
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        sn = 'Rec-COB' if 'Rec-COB' in wb.sheetnames else ('Rec_COB' if 'Rec_COB' in wb.sheetnames else None)
        if not sn:
            continue
        for r in wb[sn].iter_rows(values_only=True):
            if r and str(r[0]).strip().upper() == 'CHINA' and len(r) > 5:
                rec[m.group(1)] = (r[1], r[5]); break
    # 2) CSV（同结构，填 xlsx 没有的季度，如 USCIS 改 CSV 命名的那些）
    for f in sorted(glob.glob(os.path.join(USCIS, "I140_FY*_Q*_rec_cob.csv"))):
        m = re.search(r'I140_(FY\d+_Q\d+)', os.path.basename(f))
        if not m or m.group(1) in rec:
            continue
        v = _csv_china_cols(f)
        if v and len(v) > 5 and v[1] is not None:
            rec[m.group(1)] = (v[1], v[5])
    # 3) supplement.json（社区补充，填以上都没有的季度）
    sup = os.path.join(ROOT, "data", "i140_china_receipts_supplement.json")
    if os.path.exists(sup):
        import json
        with open(sup, encoding="utf-8") as fh:
            q = json.load(fh).get("quarters", {})
        for tag, v in q.items():
            if tag not in rec:
                rec[tag] = (v.get("eb1a"), v.get("niw"))
    return [(tag, e, n) for tag, (e, n) in sorted(rec.items())]


def adjudication_rates():
    """EB-1A 逐季获批率(全球口径,人工补录)。返回 {tag: approved_pct}。"""
    path = os.path.join(ROOT, "data", "i140_case_status_supplement.json")
    if not os.path.exists(path):
        return {}, {}
    import json
    with open(path, encoding="utf-8") as fh:
        d = json.load(fh)
    rates = {k: v["approved_pct"] for k, v in d.get("eb1a_adjudication_rate", {}).items()
             if not k.startswith("_")}
    return rates, d.get("quarters", {})


def main():
    lines = []
    def p(s): print(s); lines.append(s)

    p("## China EB1 待签存量(approved awaiting visa) —— 申请潮/退潮的领先指标")
    stock = awaiting_stock()
    prev = None
    for asof, c in stock:
        delta = f"（环比 {c-prev:+d}）" if prev is not None else ""
        p(f"  {asof}  存量 = {c:,}{delta}")
        prev = c
    if len(stock) >= 3:
        recent = stock[-1][1] - stock[-2][1]
        earlier = stock[-2][1] - stock[-3][1]
        trend = "📈 仍在变厚" if recent > earlier else ("📉 增速放缓/退潮" if recent < earlier else "持平")
        p(f"  → 最近一期环比 {recent:+d} vs 上期 {earlier:+d}：{trend}")

    p("\n## China I-140 收件(Rec-COB，单季值) —— EB-1A / NIW 申请潮")
    rs = receipts()
    for tag, eb1a, niw in rs:
        p(f"  {tag}: EB-1A {eb1a}   NIW {niw}")
    # supplement 缺值时 eb1a 可能为 None——过滤后再算峰值/趋势，否则 max()/除法直接崩、
    # 整段摘要因 workflow 的 `|| true` 静默消失
    valid = [(tag, e) for tag, e, _ in rs if isinstance(e, (int, float))]
    if len(valid) >= 2:
        e1, pk = valid[-1][1], max(e for _, e in valid)
        if pk > 0:
            trend = '退潮' if e1 < pk * 0.95 else ('见顶' if e1 < pk else '升温中')
            p(f"  → EB-1A 峰值 {pk} → 最新 {e1}（{trend}，较峰 {round((e1/pk-1)*100)}%）;"
              " 每文件为单季值。流入缩 = 未来 PD 身后堆积变小。")

    # 有效流入 = 收件 × 获批率。只看收件会严重低估退潮幅度——EB-1A 获批率已从
    # FY2025Q1 的 74.8% 一路跌到 FY2026Q2 的 41.7%,同一批收件真正能进排队的人腰斩。
    rates, quarters = adjudication_rates()
    if rates and valid:
        p("\n## EB-1A 有效流入(= 中国收件 × 获批率) —— 真实进入排队的人")
        p("  注:获批率为【全球】口径(USCIS 季报),中国审查更严、实际可能更低,故本栏属上界。")
        eff = []
        for tag, e in valid:
            r = rates.get(tag)
            if r is None:
                continue
            eff.append((tag, e * r / 100))
            p(f"  {tag}: 收件 {e} × 获批率 {r}% → 有效流入 ≈ {round(e * r / 100)}")
        if len(eff) >= 2:
            pk_e = max(v for _, v in eff)
            last_e = eff[-1][1]
            drop_recv = round((valid[-1][1] / max(x for _, x in valid) - 1) * 100)
            drop_eff = round((last_e / pk_e - 1) * 100)
            p(f"  → 有效流入 峰值 {round(pk_e)} → 最新 {round(last_e)}（较峰 {drop_eff}%）；"
              f"仅看收件为 {drop_recv}%——获批率下滑使真实退潮幅度约为收件口径的 "
              f"{abs(drop_eff)/max(abs(drop_recv),1):.1f} 倍。")

    # 待审积压:2024+ PD「需求墙」目前唯一的实测参照(I-485 库存该段仍全为 0)
    q_latest = max(quarters) if quarters else None
    if q_latest:
        e11 = quarters[q_latest].get("E11_EB1A", {})
        pend = e11.get("pending")
        # 份额必须用【同季】中国收件 ÷ 同季全球收件。跨季相除会系统性偏高
        # (全球收件逐季下滑而中国分子不变),实测差 26.3% vs 22.5%、约 400 人。
        cn_by_tag = dict(valid)
        common = sorted(t for t in quarters if t in cn_by_tag
                        and quarters[t].get("E11_EB1A", {}).get("received"))
        if pend and common:
            tag = common[-1]
            recv_w = quarters[tag]["E11_EB1A"]["received"]
            share = cn_by_tag[tag] / recv_w
            r_last = rates.get(max(rates)) if rates else None
            p(f"\n## EB-1A 待审积压({q_latest}，全球 {pend:,} 件) —— 需求墙的实测参照")
            p(f"  中国份额(同季 {tag}): {cn_by_tag[tag]} ÷ {recv_w:,} = {share*100:.1f}%"
              f" → 中国待审积压 ≈ {round(pend*share):,} 件")
            if r_last:
                p(f"  按最新获批率 {r_last}% → 未来将进入排队 ≈ {round(pend*share*r_last/100):,} 人(主申)")
            p("  注:份额按单季收件比估算,占比逐季波动,仅供量级参考;"
              "这批 PD 多在 2024+,主要影响 cutoff 进入 2024 之后的推进速度。")

    sp = os.environ.get("GITHUB_STEP_SUMMARY")
    if sp:
        with open(sp, "a", encoding="utf-8") as fh:
            fh.write("\n".join(lines) + "\n")


if __name__ == "__main__":
    main()
