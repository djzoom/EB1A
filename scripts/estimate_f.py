#!/usr/bin/env python3
"""估计 f —— 待签池里排在你 PD 之前的比例 → 推出"排到"时间区间。

第一性模型(单位一致,见 README/讨论):
    T = (POOL × family × f) / Supply_persons_per_year
其中 POOL=I-140 待签(principals)、family=家庭系数、Supply=年签发(persons)。

f = 池中 PD ≤ 你的 PD 的比例 = 池 PD 累积分布在你 PD 处的值。
EB-1 无 PERM ⇒ 池 PD = I-140 收件日 ⇒ 池 PD 分布 ≈ 收件时间分布。
但 USCIS 不公布池的逐 PD 分布，故 f 只能给"区间"，并随 cutoff 推进逐月收敛到真值。

数据(自动抓取,data/raw/uscis):
  - I140_I360_I526_Approved_*  → 池总量(China EB1 awaiting)
  - I140_FY*_Q*  Rec-COB       → China EB-1 年收件(验证流入是否平稳)
  - index.html                 → 当前 cutoff(表A)

用法: python scripts/estimate_f.py [--pd 2024-06-10] [--supply 3662] [--family 1.9]
"""
import argparse, glob, os, re
from datetime import date
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
USCIS = os.path.join(ROOT, "data/raw/uscis")
INDEX = os.path.join(ROOT, "index.html")
MO = {m: i + 1 for i, m in enumerate(
    ['January','February','March','April','May','June','July','August','September','October','November','December'])}


def latest_pool():
    """最新 China EB1 awaiting-visa 存量 (principals) 及其 as-of 月。"""
    best = None
    for f in glob.glob(os.path.join(USCIS, "I140_I360_I526_Approved_*.xlsx")):
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
        asof = None
        for r in rows[:5]:
            if r and r[0] and 'As of' in str(r[0]):
                m = re.search(r'As of (\w+)\s+(\d{4})', str(r[0]))
                if m: asof = date(int(m.group(2)), MO[m.group(1)], 1)
        china = next((r[1] for r in rows if r and str(r[0]).strip() == 'China'), None)
        if asof and isinstance(china, int) and (best is None or asof > best[0]):
            best = (asof, china)
    return best


def annual_receipts():
    """China EB-1(A+B+C)年收件，用最新的整年 Rec-COB。返回 (label, total)。"""
    out = []
    for f in sorted(glob.glob(os.path.join(USCIS, "I140_FY*_Q*.xlsx"))):
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        sn = 'Rec-COB' if 'Rec-COB' in wb.sheetnames else ('Rec_COB' if 'Rec_COB' in wb.sheetnames else None)
        if not sn: continue
        for r in wb[sn].iter_rows(values_only=True):
            if r and str(r[0]).strip().upper() == 'CHINA':
                tag = re.search(r'(FY\d+_Q\d+)', os.path.basename(f)).group(1)
                out.append((tag, (r[1] or 0) + (r[2] or 0) + (r[3] or 0)))
                break
    return out


def cutoff_from_index():
    s = open(INDEX, encoding="utf-8").read()
    m = re.search(r"'EB-1A':\s*\{\s*'CN':\s*\{\s*A:\s*'([0-9-]+)'", s)
    return date.fromisoformat(m.group(1)) if m else date(2023, 6, 1)


def months(a, b):
    return (b.year - a.year) * 12 + (b.month - a.month) + (b.day - a.day) / 30.4


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pd", default="2024-06-10")
    ap.add_argument("--supply", type=float, default=3662.0)  # persons/年(回测锚定)
    ap.add_argument("--family", type=float, default=1.9)
    args = ap.parse_args()
    your_pd = date.fromisoformat(args.pd)
    today = date.today()

    pool = latest_pool(); cutoff = cutoff_from_index(); rec = annual_receipts()
    print(f"输入: 池(China EB1 awaiting)={pool[1]:,} (as of {pool[0]});  当前 cutoff={cutoff};  你的 PD={your_pd}")
    print(f"      Supply={args.supply:.0f} 人/年;  家庭系数={args.family}")
    print("      年收件(China EB-1, 验证流入平稳): " + ", ".join(f"{t}:{v}" for t, v in rec))

    POOL = pool[1]
    # 池 PD 跨度 = [cutoff, 今天](I-140 至今才能收件，PD 不超过今天)
    span = months(cutoff, today)
    window = months(cutoff, your_pd)
    base = window / span  # 均匀假设
    print(f"\n池 PD 跨度 ≈ [{cutoff} → {today}] = {span:.0f} 月;  你前方窗口 [{cutoff} → {your_pd}] = {window:.0f} 月")
    print(f"均匀分布 f = {window:.0f}/{span:.0f} = {base:.2f}")

    # 形状不确定性:收件温和增长→近端(高PD)略重→你窗口占比略低;领事滞后→略平→略高。带 ±20%
    scen = {"近端加权(偏乐观)": base * 0.8, "均匀(中枢)": base, "前端加权(偏保守)": min(1.0, base * 1.2)}
    print(f"\n{'情形':<18}{'f':>6}{'前方厚度(人)':>14}{'T(年)':>8}{'排到':>10}")
    for name, f in scen.items():
        thick = POOL * args.family * f
        T = thick / args.supply
        m_total = today.month - 1 + round(T * 12)
        eta = f"{today.year + m_total // 12}-{m_total % 12 + 1:02d}"
        print(f"{name:<18}{f:>6.2f}{round(thick):>14,}{T:>8.2f}{eta:>10}")

    # 数据自洽校验:收件率 vs 池总量
    if rec:
        rate = max(v for _, v in rec)  # 取最高年化作上界估计
        implied_span = POOL / (rate / 12)
        print(f"\n[自洽校验] 池/收件率 ⇒ 跨度≈{implied_span:.0f}月，而 [cutoff→今天]={span:.0f}月。"
              f"\n  若前者明显更大 ⇒ 池含早于 cutoff 的领事积压(都排你前面)⇒ f 偏高、排到更晚；"
              f"\n  这是 f 最大的不确定来源，须待 cutoff 进入 2024 PD 后由 I-485 库存实测收敛。")


if __name__ == "__main__":
    main()
