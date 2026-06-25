#!/usr/bin/env python3
"""回测校准 v2 —— 时点匹配快照（Charlie calibration）。

相比 v1：每个回测步用「当时的」I-485 库存快照读密度，消除快照偏差
（同一 PD 的库存随 DFF 放开而累积、随裁定而消耗，不能用单一晚快照代替历史）。

输入(真实)：
  1) cutoff 轨迹  : index.html HISTORY(表A)
  2) PD 月密度    : data/raw/uscis/I485_Pending_Inventory_*.xlsx 多份快照(按 as-of 时点匹配)
  3) 未来需求池   : I-140 China EB1 awaiting = 13,598

输出：
  A. 用时点匹配密度，拟合近窗口有效年供给，与模型/官方对照；
  B. rolling-holdout(不泄露未来快照)：预测误差；
  C. 未来需求膨胀因子。
"""
import re, os, glob
from datetime import date, timedelta
from collections import defaultdict
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INDEX = os.path.join(ROOT, "index.html")
SNAP_GLOB = os.path.join(ROOT, "data/raw/uscis/I485_Pending_Inventory_*.xlsx")
I140_CHINA_EB1 = 13598
MO = ['January','February','March','April','May','June','July','August','September','October','November','December']
MI = {m: i + 1 for i, m in enumerate(MO)}
W = [0.338,0.183,0.139,0.110,0.057,0.037,0.047,0.017,0.022,0.026,0.020,0.004]


def parse_history():
    s = open(INDEX, encoding="utf-8").read()
    m = re.search(r"var HISTORY = \[(.*?)\]\.map", s, re.S)
    pts = re.findall(r"\['(\d{4}-\d{2}-\d{2})','(\d{4}-\d{2}-\d{2})'\]", m.group(1))
    return [(date.fromisoformat(a), date.fromisoformat(b)) for a, b in pts]


def load_snapshots():
    snaps = []
    for f in glob.glob(SNAP_GLOB):
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        if 'China' not in wb.sheetnames:
            continue
        ws = wb['China']; rows = list(ws.iter_rows(values_only=True)); asof = None
        for r in rows[:4]:
            if r[0] and 'As of' in str(r[0]):
                mm = re.search(r'As of (\w+) (\d+), (\d+)', str(r[0]))
                if mm: asof = date(int(mm.group(3)), MI[mm.group(1)], int(mm.group(2)))
        if not asof:
            continue
        hdr = rows[3]; yc = {h.split('- ')[-1]: j for j, h in enumerate(hdr) if isinstance(h, str) and 'Priority Date Year - 2' in h}
        g = defaultdict(int)
        for r in rows[4:]:
            pref, st, mon = r[1], r[2], r[3]
            if not pref or 'EB1' not in str(pref) or mon not in MI:
                continue
            for y, j in yc.items():
                v = r[j]
                if isinstance(v, (int, float)): g[(int(y), MI[mon])] += v
        snaps.append((asof, dict(g)))
    snaps.sort(key=lambda x: x[0])
    return snaps


def grid_asof(snaps, t, max_asof):
    cap = min(t, max_asof)
    pick = None
    for asof, g in snaps:
        if asof <= cap: pick = g
    return pick or (snaps[0][1] if snaps else {})


def dpd(g, cut):
    """该 PD 月真实库存/30.4 = 人/PD天；为 0 时回退到该快照非零均值。"""
    v = g.get((cut.year, cut.month), 0)
    if v == 0:
        nz = [x for x in g.values() if x > 0]
        v = (sum(nz) / len(nz)) if nz else 400
    return v / 30.4


def add_days(d, n): return d + timedelta(days=n)
def nm(d): return date(d.year + d.month // 12, d.month % 12 + 1, 15)


def simulate(snaps, start_bd, start_cut, end_bd, annual, max_asof):
    cut = start_cut; cur = date(start_bd.year, start_bd.month, 15); end = date(end_bd.year, end_bd.month, 15)
    while cur < end:
        g = grid_asof(snaps, cur, max_asof)
        adv = annual * W[(cur.month - 10) % 12] / dpd(g, cut)
        cut = add_days(cut, round(adv)); cur = nm(cur)
    return cut


def fit(snaps, start, end, target, max_asof):
    lo, hi = 300, 20000
    for _ in range(44):
        mid = (lo + hi) / 2
        got = simulate(snaps, start[0], start[1], end[0], mid, max_asof)
        if got < target: lo = mid
        else: hi = mid
    return (lo + hi) / 2


def main():
    hist = parse_history(); snaps = load_snapshots()
    if len(hist) < 8 or not snaps:
        print(f"数据不足(history={len(hist)}, snapshots={len(snaps)})，无法回测校准。"); return
    win = hist[-8:]
    print("快照 as-of:", ", ".join(str(a) for a, _ in snaps))

    print("\n" + "=" * 60 + "\nA. 时点匹配密度 → 拟合近窗口有效年供给\n" + "=" * 60)
    av = fit(snaps, win[0], win[-1], win[-1][1], max_asof=win[-1][0])
    (b0, c0), (b1, c1) = win[0], win[-1]
    print(f"窗口 {b0}→{b1}: cutoff {c0}→{c1}  (+{(c1-c0).days}天)")
    print(f"时点匹配拟合的有效年供给 = {av:.0f} 人/年")
    print("  对照: 官方实际 ~3500–4900；模型 totalVisas≈3803")

    print("\n" + "=" * 60 + "\nB. Rolling-holdout（不泄露未来快照）\n" + "=" * 60)
    for split_i in (3, 5):
        te = win[split_i]
        av_t = fit(snaps, win[0], te, te[1], max_asof=te[0])
        print(f"\n训练 {win[0][0]}→{te[0]}  拟合年供给={av_t:.0f}  (只用 ≤{te[0]} 的快照)")
        errs = []
        for bd, actual in win[split_i + 1:]:
            pred = simulate(snaps, te[0], te[1], bd, av_t, max_asof=te[0])
            e = (pred - actual).days; errs.append(abs(e))
            print(f"  预测 {bd}: 模型 {pred} 实际 {actual} 误差 {e:+d} 天")
        if errs: print(f"  平均绝对误差 = {sum(errs)/len(errs):.0f} 天")

    print("\n" + "=" * 60 + "\nC. 未来需求膨胀因子\n" + "=" * 60)
    latest = snaps[-1][1]; tot = sum(latest.values())
    print(f"最新 I-485 库存={tot}; I-140 awaiting={I140_CHINA_EB1} → 未来(2024+PD)需求 ×{I140_CHINA_EB1/tot:.2f} 风险")


if __name__ == "__main__":
    main()
